import os
import streamlit as st
import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO
import streamlit.components.v1 as components

# --- 1. إعدادات الصفحة والواجهة العامة ---
st.set_page_config(
    page_title="الشركة العربية للأدوية | المنظومة الرقمية",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- 2. إدارة حالة الجلسة (Session State) لتسجيل الدخول ---
if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False
if 'user_fullname' not in st.session_state:
    st.session_state['user_fullname'] = ''

# --- 3. التنسيقات البصرية الهيكلية (CSS) ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;800&display=swap');

    html, body, [class*="css"], div, span, label {
        font-family: 'Cairo', sans-serif !important;
        direction: rtl;
        text-align: right;
    }

    .stApp {
        background-color: #f8fafc;
    }

    /* بطاقات الأقسام */
    .card-box {
        background-color: #ffffff;
        padding: 24px;
        border-radius: 16px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.04);
        margin-bottom: 20px;
    }

    /* عناوين الأقسام */
    .section-title {
        color: #0d5c75;
        font-size: 18px;
        font-weight: 700;
        margin-bottom: 15px;
        display: flex;
        align-items: center;
        gap: 8px;
        border-bottom: 2px solid #f1f5f9;
        padding-bottom: 8px;
    }

    /* تخصيص أزرار التحكم الرئيسية */
    div.stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #0d5c75 0%, #053b4c 100%) !important;
        color: white !important;
        font-size: 17px !important;
        font-weight: 700 !important;
        padding: 10px 20px !important;
        border-radius: 10px !important;
        border: none !important;
        box-shadow: 0 4px 14px rgba(13, 92, 117, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    
    div.stButton > button:hover {
        background: linear-gradient(135deg, #d4a373 0%, #b88656 100%) !important;
        box-shadow: 0 6px 18px rgba(212, 163, 115, 0.3) !important;
        transform: translateY(-2px);
    }

    /* زر التحميل الاخضر */
    div.stDownloadButton > button {
        width: 100%;
        background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
        color: white !important;
        font-size: 18px !important;
        font-weight: 700 !important;
        padding: 12px 24px !important;
        border-radius: 10px !important;
        border: none !important;
        box-shadow: 0 4px 14px rgba(16, 185, 129, 0.25) !important;
        transition: all 0.3s ease !important;
    }

    div.stDownloadButton > button:hover {
        background: linear-gradient(135deg, #059669 0%, #047857 100%) !important;
        transform: translateY(-2px);
    }

    /* المقاييس Metrics */
    [data-testid="stMetricValue"] {
        font-size: 24px !important;
        font-weight: 800 !important;
        color: #0d5c75 !important;
    }

    [data-testid="stMetricBackground"] {
        background-color: #ffffff !important;
        border-radius: 12px !important;
        padding: 12px !important;
        border: 1px solid #e2e8f0 !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.02) !important;
    }
</style>
""", unsafe_allow_html=True)


# --- 4. شاشة تسجيل الدخول ---
def show_login_page():
    st.markdown("""
    <div style="text-align: center; margin-top: 30px; margin-bottom: 20px;">
        <h1 style="color: #0d5c75; font-size: 32px; font-weight: 800;">الشركة العربية للأدوية المحدودة</h1>
        <h4 style="color: #d4a373; font-weight: 700;">Arab Pharmaceuticals Co. Ltd.</h4>
        <p style="color: #64748b; font-size: 15px;">بوابة الدخول للمنظومة الإدارية الرقمية</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown('<div class="card-box">', unsafe_allow_html=True)
        st.markdown('<h3 style="text-align:center; color:#0d5c75; font-weight:700; margin-bottom:20px;">🔐 تسجيل الدخول</h3>', unsafe_allow_html=True)
        
        username = st.text_input("اسم المستخدم", placeholder="أدخل اسم المستخدم")
        password = st.text_input("كلمة المرور", type="password", placeholder="أدخل كلمة المرور")
        
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("دخول المنظومة"):
            # بيانات تجريبية لتسجيل الدخول (يمكن تعديلها أو ربطها بقاعدة بيانات)
            if username.strip() == "apc" and password == "123456":
                st.session_state['logged_in'] = True
                st.session_state['user_fullname'] = "د. شفيق الرامسي"
                st.success("تم تسجيل الدخول بنجاح!")
                st.rerun()
            elif username.strip() != "" and password != "":
                # السماح بالدخول لأي اسم مستخدم مراد لتسهيل التجربة
                st.session_state['logged_in'] = True
                st.session_state['user_fullname'] = username
                st.success("تم تسجيل الدخول بنجاح!")
                st.rerun()
            else:
                st.error("يرجى إدخال اسم المستخدم وكلمة المرور بشكل صحيح.")
        st.markdown('</div>', unsafe_allow_html=True)


# --- 5. شفرة صفحة التفقيط HTML/JS مدمجة ---
HTML_TAFQEET_PAGE = """
<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>أداة تحويل الأرقام إلى نصوص - APC</title>
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;800&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {
            --primary: #057586;
            --primary-dark: #035460;
            --primary-light: #0FA7B7;
            --secondary: #1a3a5f;
            --bg-color: #f8fafc;
            --card-bg: #ffffff;
            --text-main: #333333;
            --text-muted: #6c757d;
            --border-color: #e2e8f0;
            --success: #28a745;
            --shadow-sm: 0 2px 4px rgba(0,0,0,0.05);
            --shadow-md: 0 10px 20px rgba(0,0,0,0.08);
            --shadow-lg: 0 15px 35px rgba(0,0,0,0.1);
        }

        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
            font-family: 'Cairo', sans-serif;
        }

        body {
            background-color: var(--bg-color);
            color: var(--text-main);
            display: flex;
            flex-direction: column;
            align-items: center;
            padding: 10px;
        }

        .container {
            background-color: var(--card-bg);
            border-radius: 20px;
            box-shadow: var(--shadow-md);
            width: 100%;
            max-width: 1100px;
            overflow: hidden;
            margin-bottom: 20px;
            border: 1px solid var(--border-color);
        }

        .header {
            background: linear-gradient(135deg, var(--secondary), var(--primary));
            color: white;
            padding: 25px 30px;
            display: flex;
            align-items: center;
            gap: 20px;
        }

        .header-text h1 {
            font-size: 26px;
            font-weight: 700;
            margin-bottom: 5px;
        }

        .header-text p {
            font-size: 15px;
            opacity: 0.9;
        }
        
        .tabs {
            display: flex;
            background-color: #f1f5f9;
            border-bottom: 1px solid var(--border-color);
        }

        .tab {
            flex: 1;
            padding: 16px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s ease;
            font-weight: 700;
            font-size: 17px;
            color: var(--text-muted);
            border-bottom: 3px solid transparent;
        }

        .tab.active {
            background-color: var(--card-bg);
            border-bottom: 3px solid var(--primary);
            color: var(--primary);
        }
        
        .tab-content {
            display: none;
            padding: 30px;
        }

        .tab-content.active {
            display: block;
            animation: fadeIn 0.4s ease-in-out;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }

        .layout-grid {
            display: grid;
            grid-template-columns: 1.2fr 1fr;
            gap: 30px;
            align-items: start;
        }

        .input-group {
            margin-bottom: 20px;
        }

        .input-group label {
            display: block;
            margin-bottom: 8px;
            font-weight: 700;
            color: var(--secondary);
            font-size: 15px;
        }

        .input-container {
            position: relative;
            display: flex;
            align-items: center;
        }

        .input-container span {
            position: absolute;
            left: 15px;
            color: var(--primary);
            font-weight: 800;
            font-size: 18px;
            user-select: none;
        }

        input[type="text"] {
            width: 100%;
            padding: 14px 15px 14px 45px;
            border: 2px solid var(--border-color);
            border-radius: 12px;
            font-size: 17px;
            font-weight: 600;
            color: var(--secondary);
            background-color: #f8fafc;
            text-align: right;
            transition: all 0.3s;
        }

        input[type="text"]:focus {
            border-color: var(--primary-light);
            background-color: #ffffff;
            outline: none;
            box-shadow: 0 0 0 4px rgba(15, 167, 183, 0.1);
        }
        
        .options-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(110px, 1fr));
            gap: 12px;
            margin-bottom: 20px;
        }

        .option-card {
            padding: 12px;
            border: 2px solid var(--border-color);
            border-radius: 12px;
            text-align: center;
            cursor: pointer;
            transition: all 0.2s ease;
            background: #ffffff;
            font-weight: 600;
            color: var(--text-muted);
        }

        .option-card.selected {
            border-color: var(--primary);
            background-color: #f0fdfa;
            color: var(--primary-dark);
            box-shadow: 0 0 0 1px var(--primary);
        }
        
        .cents-group {
            display: none;
            animation: fadeIn 0.3s ease;
        }

        .cents-group.visible {
            display: block;
        }

        .button-group {
            display: flex;
            gap: 12px;
            margin-bottom: 25px;
        }

        button.action-btn {
            background: linear-gradient(135deg, var(--primary), var(--primary-light));
            color: white;
            border: none;
            padding: 14px;
            border-radius: 12px;
            cursor: pointer;
            font-size: 17px;
            font-weight: 700;
            flex: 2;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            box-shadow: 0 4px 15px rgba(5, 117, 134, 0.25);
        }

        button.action-btn:hover {
            background: linear-gradient(135deg, var(--primary-dark), var(--primary));
            transform: translateY(-2px);
        }

        button.secondary {
            flex: 1;
            background: #f1f5f9;
            color: var(--text-muted);
            box-shadow: none;
            border: 1px solid var(--border-color);
        }

        button.secondary:hover {
            background: #e2e8f0;
            color: #ef4444;
            border-color: #ef4444;
        }

        .result-box {
            background: linear-gradient(to left, #f8fafc, #ffffff);
            border: 1px solid var(--border-color);
            border-right: 5px solid var(--primary);
            border-radius: 15px;
            padding: 20px;
            position: relative;
        }

        .result-box h3 {
            color: var(--secondary);
            font-size: 16px;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .result-text {
            font-size: 20px;
            line-height: 1.6;
            color: var(--primary-dark);
            font-weight: 700;
            min-height: 35px;
        }

        .copy-result-btn {
            margin-top: 15px;
            padding: 8px 16px;
            font-size: 13px;
            background: var(--secondary);
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

        .side-card {
            background-color: #f8fafc;
            border: 1px solid var(--border-color);
            border-radius: 15px;
            padding: 18px;
            margin-bottom: 20px;
        }

        .side-card h3 {
            color: var(--secondary);
            font-size: 16px;
            margin-bottom: 12px;
            padding-bottom: 8px;
            border-bottom: 2px solid var(--border-color);
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .text-item {
            padding: 10px 12px;
            background: #ffffff;
            border: 1px solid var(--border-color);
            border-radius: 8px;
            margin-bottom: 8px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .text-content {
            font-size: 14px;
            font-weight: 600;
            color: var(--text-main);
        }

        .copy-btn {
            background: transparent;
            color: var(--primary);
            border: 1px solid var(--primary);
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: 700;
            cursor: pointer;
        }

        .copy-btn:hover {
            background: var(--primary);
            color: white;
        }

        .history-list {
            max-height: 200px;
            overflow-y: auto;
        }

        .history-item {
            padding: 10px 0;
            border-bottom: 1px dashed var(--border-color);
        }

        .history-value {
            font-size: 14px;
            font-weight: 700;
            color: var(--secondary);
        }

        .history-time {
            font-size: 11px;
            color: var(--text-muted);
        }

        @media (max-width: 800px) {
            .layout-grid {
                grid-template-columns: 1fr;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-text">
                <h1>أداة تحويل الأرقام والهدايا إلى نصوص</h1>
                <p>نظام احترافي دقيق لتحويل المبالغ والأرقام إلى نصوص عربية معتمدة</p>
            </div>
        </div>
        
        <div class="tabs">
            <div class="tab active" data-tab="currency"><i class="fas fa-coins"></i> تحويل العملات</div>
            <div class="tab" data-tab="gifts"><i class="fas fa-gift"></i> تحويل الهدايا</div>
        </div>
        
        <div class="tab-content active" id="currency-tab">
            <div class="layout-grid">
                <div class="main-panel">
                    <div class="input-group">
                        <label for="numberInput">أدخل المبلغ (بالأرقام):</label>
                        <div class="input-container">
                            <span>﷼</span>
                            <input type="text" id="numberInput" placeholder="مثال: 20,000" inputmode="numeric">
                        </div>
                    </div>
                    
                    <div class="input-group cents-group" id="centsGroup">
                        <label for="centsInput">أدخل السنتات (اختياري):</label>
                        <div class="input-container">
                            <span>¢</span>
                            <input type="text" id="centsInput" placeholder="مثال: 99" inputmode="numeric" maxlength="2">
                        </div>
                    </div>
                    
                    <div class="input-group">
                        <label>اختر العملة:</label>
                        <div class="options-grid">
                            <div class="option-card selected" id="currencyYER">
                                <i class="fas fa-money-bill-wave"></i> ريال يمني
                            </div>
                            <div class="option-card" id="currencyUSD">
                                <i class="fas fa-dollar-sign"></i> دولار أمريكي
                            </div>
                        </div>
                    </div>
                    
                    <div class="button-group">
                        <button class="action-btn" onclick="convertNumber()">
                            <i class="fas fa-magic"></i> تحويل إلى نص
                        </button>
                        <button class="action-btn secondary" onclick="clearAll()">
                            <i class="fas fa-trash-alt"></i> تفريغ
                        </button>
                    </div>
                    
                    <div class="result-box">
                        <h3><i class="fas fa-check-circle"></i> النتيجة النهائية:</h3>
                        <p id="resultOutput" class="result-text">ستظهر النتيجة هنا...</p>
                        <button class="copy-result-btn" onclick="copyToClipboard(this, 'resultOutput')">
                            <i class="fas fa-copy"></i> نسخ النتيجة
                        </button>
                    </div>
                </div>

                <div class="side-panel">
                    <div class="side-card">
                        <h3><i class="fas fa-clipboard-list"></i> ملاحظات جاهزة</h3>
                        <div class="text-item">
                            <div class="text-content">هدية نقدية وعينية بحسب الاتفاقية</div>
                            <button class="copy-btn" onclick="copyText(this)">نسخ</button>
                        </div>
                        <div class="text-item">
                            <div class="text-content">هدية نقدية بحسب الاتفاقية</div>
                            <button class="copy-btn" onclick="copyText(this)">نسخ</button>
                        </div>
                        <div class="text-item">
                            <div class="text-content">هدية عينية بحسب الاستثناء</div>
                            <button class="copy-btn" onclick="copyText(this)">نسخ</button>
                        </div>
                        <div class="text-item">
                            <div class="text-content">هدية شرط الوصول الى الكمية</div>
                            <button class="copy-btn" onclick="copyText(this)">نسخ</button>
                        </div>
                    </div>

                    <div class="side-card">
                        <h3><i class="fas fa-history"></i> السجل الأخير</h3>
                        <div id="historyList" class="history-list"></div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="tab-content" id="gifts-tab">
            <div class="layout-grid">
                <div class="main-panel">
                    <div class="input-group">
                        <label for="giftNumberInput">أدخل العدد (بالأرقام):</label>
                        <div class="input-container">
                            <span>#</span>
                            <input type="text" id="giftNumberInput" placeholder="مثال: 25" inputmode="numeric">
                        </div>
                    </div>
                    
                    <div class="input-group">
                        <label>اختر نوع الهدية (الوحدة):</label>
                        <div class="options-grid">
                            <div class="option-card selected" data-gift="vial">
                                <i class="fas fa-vial"></i> فيالة
                            </div>
                            <div class="option-card" data-gift="pack">
                                <i class="fas fa-box"></i> باكت
                            </div>
                            <div class="option-card" data-gift="box">
                                <i class="fas fa-box-open"></i> علبة
                            </div>
                        </div>
                    </div>
                    
                    <div class="button-group">
                        <button class="action-btn" onclick="convertGift()">
                            <i class="fas fa-magic"></i> تحويل إلى نص
                        </button>
                        <button class="action-btn secondary" onclick="clearGifts()">
                            <i class="fas fa-trash-alt"></i> تفريغ
                        </button>
                    </div>
                    
                    <div class="result-box">
                        <h3><i class="fas fa-check-circle"></i> النتيجة النهائية:</h3>
                        <p id="giftResultOutput" class="result-text">ستظهر النتيجة هنا...</p>
                        <button class="copy-result-btn" onclick="copyToClipboard(this, 'giftResultOutput')">
                            <i class="fas fa-copy"></i> نسخ النتيجة
                        </button>
                    </div>
                </div>

                <div class="side-panel">
                    <div class="side-card">
                        <h3><i class="fas fa-clipboard-list"></i> ملاحظات جاهزة</h3>
                        <div class="text-item">
                            <div class="text-content">هدية نقدية وعينية بحسب الاتفاقية</div>
                            <button class="copy-btn" onclick="copyText(this)">نسخ</button>
                        </div>
                        <div class="text-item">
                            <div class="text-content">هدية عينية بحسب الاتفاقية</div>
                            <button class="copy-btn" onclick="copyText(this)">نسخ</button>
                        </div>
                    </div>

                    <div class="side-card">
                        <h3><i class="fas fa-history"></i> سجل الهدايا</h3>
                        <div id="giftHistoryList" class="history-list"></div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script>
        let conversionHistory = [];
        let giftConversionHistory = [];
        let selectedCurrency = 'YER'; 
        let selectedGift = 'vial'; 
        
        document.querySelectorAll('.tab').forEach(tab => {
            tab.addEventListener('click', function() {
                document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
                document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
                this.classList.add('active');
                document.getElementById(this.dataset.tab + '-tab').classList.add('active');
            });
        });
        
        document.getElementById('currencyYER').addEventListener('click', () => selectCurrency('YER'));
        document.getElementById('currencyUSD').addEventListener('click', () => selectCurrency('USD'));
        
        function selectCurrency(currency) {
            selectedCurrency = currency;
            document.getElementById('currencyYER').classList.toggle('selected', currency === 'YER');
            document.getElementById('currencyUSD').classList.toggle('selected', currency === 'USD');
            document.querySelector('#currency-tab .input-container span').textContent = currency === 'YER' ? '﷼' : '$';
            
            const centsGroup = document.getElementById('centsGroup');
            if (currency === 'USD') {
                centsGroup.classList.add('visible');
            } else {
                centsGroup.classList.remove('visible');
                document.getElementById('centsInput').value = '';
            }
        }
        
        document.querySelectorAll('#gifts-tab .option-card').forEach(option => {
            option.addEventListener('click', function() {
                document.querySelectorAll('#gifts-tab .option-card').forEach(o => o.classList.remove('selected'));
                this.classList.add('selected');
                selectedGift = this.dataset.gift;
            });
        });
        
        function formatNumberInput(e) {
            let input = e.target.value.replace(/[^\\d]/g, '');
            if (input.length > 3) {
                input = input.replace(/\\B(?=(\\d{3})+(?!\\d))/g, ",");
            }
            e.target.value = input;
        }

        document.getElementById('numberInput').addEventListener('input', formatNumberInput);
        document.getElementById('giftNumberInput').addEventListener('input', formatNumberInput);

        function convertNumber() {
            const input = document.getElementById('numberInput').value;
            const centsInput = document.getElementById('centsInput').value;
            const resultElement = document.getElementById('resultOutput');
            
            let cleanInput = input.replace(/,/g, '');
            let cleanCents = centsInput.replace(/,/g, '');
            
            if (!cleanInput) {
                resultElement.textContent = "يرجى إدخال مبلغ صحيح";
                return;
            }
            
            let number = parseInt(cleanInput);
            if (isNaN(number) || number >= 1000000000000) {
                resultElement.textContent = "الرجاء إدخال رقم مقبول أصل من تريليون";
                return;
            }
            
            let centsText = '';
            if (selectedCurrency === 'USD' && cleanCents) {
                let cents = parseInt(cleanCents);
                if (cents > 0 && cents < 100) {
                    centsText = convertNumberToArabicWords(cents) + " سنتاً";
                }
            }
            
            const resultText = convertNumberToArabicWords(number);
            const currencyText = selectedCurrency === 'YER' ? "ريال يمني" : "دولار أمريكي";
            
            let fullResult = resultText + " " + currencyText;
            if (centsText) fullResult += " و" + centsText;
            fullResult += " فقط لا غير";
            
            resultElement.textContent = fullResult;
            addToHistory(input, fullResult, conversionHistory, 'historyList');
        }
        
        function convertGift() {
            const input = document.getElementById('giftNumberInput').value;
            const resultElement = document.getElementById('giftResultOutput');
            
            let cleanInput = input.replace(/,/g, '');
            if (!cleanInput) {
                resultElement.textContent = "يرجى إدخال عدد صحيح";
                return;
            }
            
            let number = parseInt(cleanInput);
            if (isNaN(number) || number >= 1000000) {
                resultElement.textContent = "الرجاء إدخال عدد مقبول";
                return;
            }
            
            const resultText = convertNumberToArabicWords(number);
            let giftText = selectedGift === 'vial' ? (number === 1 ? 'فيالة' : number === 2 ? 'فيالتان' : (number > 2 && number <= 10) ? 'فيالات' : 'فيالة')
                        : selectedGift === 'pack' ? (number === 1 ? 'باكت' : number === 2 ? 'باكتان' : (number > 2 && number <= 10) ? 'بواكت' : 'باكت')
                        : (number === 1 ? 'علبة' : number === 2 ? 'علبتان' : (number > 2 && number <= 10) ? 'علب' : 'علبة');
            
            const fullResult = resultText + " " + giftText;
            resultElement.textContent = fullResult;
            addToHistory(input, fullResult, giftConversionHistory, 'giftHistoryList');
        }
        
        function addToHistory(input, result, historyArray, listId) {
            historyArray.unshift({ input, result, time: new Date().toLocaleTimeString('ar-YE', {hour: '2-digit', minute:'2-digit'}) });
            if (historyArray.length > 5) historyArray.pop();
            
            const historyList = document.getElementById(listId);
            historyList.innerHTML = historyArray.map(item => `
                <div class="history-item">
                    <div class="history-value">${item.input} = ${item.result}</div>
                    <div class="history-time">${item.time}</div>
                </div>
            `).join('');
        }
        
        function clearAll() {
            document.getElementById('numberInput').value = '';
            document.getElementById('centsInput').value = '';
            document.getElementById('resultOutput').textContent = 'ستظهر النتيجة هنا...';
        }
        
        function clearGifts() {
            document.getElementById('giftNumberInput').value = '';
            document.getElementById('giftResultOutput').textContent = 'ستظهر النتيجة هنا...';
        }
        
        function copyToClipboard(btnElement, targetId) {
            const text = document.getElementById(targetId).textContent;
            if (text.includes('ستظهر النتيجة')) return;
            navigator.clipboard.writeText(text).then(() => {
                btnElement.innerHTML = '<i class="fas fa-check"></i> تم النسخ';
                setTimeout(() => btnElement.innerHTML = '<i class="fas fa-copy"></i> نسخ النتيجة', 2000);
            });
        }
        
        function copyText(btnElement) {
            const text = btnElement.previousElementSibling.textContent;
            navigator.clipboard.writeText(text).then(() => {
                btnElement.innerHTML = 'تم';
                setTimeout(() => btnElement.innerHTML = 'نسخ', 2000);
            });
        }
        
        function convertNumberToArabicWords(number) {
            if (number === 0) return "صفر";
            const units = ['', 'واحد', 'اثنان', 'ثلاثة', 'أربعة', 'خمسة', 'ستة', 'سبعة', 'ثمانية', 'تسعة'];
            const teens = ['عشرة', 'أحد عشر', 'اثنا عشر', 'ثلاثة عشر', 'أربعة عشر', 'خمسة عشر', 'ستة عشر', 'سبعة عشر', 'ثمانية عشر', 'تسعة عشر'];
            const tens = ['', 'عشرة', 'عشرون', 'ثلاثون', 'أربعون', 'خمسون', 'ستون', 'سبعون', 'ثمانون', 'تسعون'];
            const hundreds = ['', 'مائة', 'مئتان', 'ثلاثمائة', 'أربعمائة', 'خمسمائة', 'ستمائة', 'سبعمائة', 'ثمانمائة', 'تسعمائة'];
            
            let result = '';
            const billions = Math.floor(number / 1000000000);
            if (billions > 0) {
                result += billions === 1 ? "مليار و" : billions === 2 ? "ملياران و" : convertNumberToArabicWords(billions) + ' مليار و';
                number %= 1000000000;
            }
            const millions = Math.floor(number / 1000000);
            if (millions > 0) {
                result += millions === 1 ? "مليون و" : millions === 2 ? "مليونان و" : convertNumberToArabicWords(millions) + ' مليون و';
                number %= 1000000;
            }
            const thousands = Math.floor(number / 1000);
            if (thousands > 0) {
                result += thousands === 1 ? "ألف و" : thousands === 2 ? "ألفان و" : convertNumberToArabicWords(thousands) + ' ألف و';
                number %= 1000;
            }
            const h = Math.floor(number / 100);
            if (h > 0) {
                result += hundreds[h] + ' و';
                number %= 100;
            }
            if (number > 0) {
                if (number < 10) result += units[number];
                else if (number < 20) result += teens[number - 10];
                else {
                    const t = Math.floor(number / 10);
                    const u = number % 10;
                    result += u === 0 ? tens[t] : units[u] + ' و' + tens[t];
                }
            }
            if (result.endsWith(' و')) result = result.slice(0, -2);
            return result;
        }
    </script>
</body>
</html>
"""


# --- 6. صفحة نظام تقارير المناديب ---
def show_reports_page():
    logo_filename = "APC logo.png"

    col_head_title, col_head_logo = st.columns([3, 1])

    with col_head_title:
        st.markdown("""
        <div style="padding-top: 10px;">
            <h1 style="color: #0d5c75; font-weight: 800; font-size: 28px; margin: 0;">الشركة العربية للأدوية المحدودة</h1>
            <h3 style="color: #d4a373; font-weight: 700; font-size: 18px; margin-top: 4px; margin-bottom: 0;">Arab Pharmaceuticals Co. Ltd.</h3>
            <p style="color: #64748b; font-size: 14px; margin-top: 6px;">نظام توليد وتقارير حركة ومتابعة زيارات المناديب المنسقة تلقائياً</p>
        </div>
        """, unsafe_allow_html=True)

    with col_head_logo:
        if os.path.exists(logo_filename):
            st.image(logo_filename, use_container_width=True)
        else:
            st.info("📊 APC Logo")

    st.markdown("<hr style='margin-top: 10px; margin-bottom: 25px; border-color: #cbd5e1;'>", unsafe_allow_html=True)

    # مدخلات المستخدم
    st.markdown('<div class="section-title">📅 إعدادات التقرير والملف الخام</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1, 2])

    with col1:
        start_date = st.date_input("تاريخ بداية التقرير", date(2026, 7, 21))

    with col2:
        end_date = st.date_input("تاريخ نهاية التقرير", date(2026, 8, 20))

    with col3:
        uploaded_file = st.file_uploader("رفع ملف الإكسل الخام (.xlsx)", type=["xlsx"])

    st.markdown("<br>", unsafe_allow_html=True)

    # زر المعالجة
    if st.button("🚀 إصدار التقرير النهائي") and uploaded_file is not None:
        with st.spinner("جاري معالجة البيانات وبناء ملف الإكسل المنسق..."):
            try:
                df_full = pd.read_excel(uploaded_file, sheet_name='Table 1', skiprows=5)
                branch_header = str(df_full.columns[0])
                branch_name = branch_header.split(':')[-1].strip() if "الفرع" in branch_header else "غير محدد"
                
                rep_data = {}
                current_rep = None
                for idx, row in df_full.iterrows():
                    rep_val = row['Unnamed: 10'] if 'Unnamed: 10' in df_full.columns else row.iloc[-1]
                    if pd.notna(rep_val) and "المندوب" not in str(rep_val):
                        current_rep = str(rep_val).strip()
                        rep_data[current_rep] = []
                        continue
                    date_col = 'Unnamed: 9' if 'Unnamed: 9' in df_full.columns else df_full.columns[-2]
                    if pd.isna(row[date_col]) or str(row[date_col]).strip() == 'التاريخ':
                        continue
                    if current_rep:
                        rep_data[current_rep].append({
                            'total_time': row.iloc[0], 'eve_end': row.iloc[1], 'eve_start': row.iloc[2],
                            'mor_end': row.iloc[3], 'mor_start': row.iloc[4], 'tot_reports': row.iloc[5],
                            'eve_reports': row.iloc[6], 'mor_reports': row.iloc[7],
                            'day_str': row.iloc[8], 'date': row[date_col]
                        })

                def clean_time(t_val):
                    if pd.isna(t_val): return ""
                    t_str = str(t_val).strip().split('.')[0]
                    dt = pd.to_datetime(t_str, format='%H:%M:%S', errors='coerce')
                    return t_str if pd.isna(dt) else dt.strftime('%I:%M:%S %p')

                def clean_total_time(t_val):
                    if pd.isna(t_val): return ""
                    t_str = str(t_val).strip().split('.')[0]
                    try:
                        parts = t_str.split(':')
                        if len(parts) == 3: return f"{int(parts[0]):02d}:{int(parts[1]):02d}:{int(parts[2]):02d}"
                    except: pass
                    return t_str

                date_range = pd.date_range(pd.to_datetime(start_date), pd.to_datetime(end_date))
                days_ar = {0: 'الإثنين', 1: 'الثلاثاء', 2: 'الأربعاء', 3: 'الخميس', 4: 'الجمعة', 5: 'السبت', 6: 'الأحد'}
                processed_reps = {}

                total_all_visits = 0
                total_all_absence = 0

                for rep, rows in rep_data.items():
                    df_rep = pd.DataFrame(rows)
                    df_rep['date'] = pd.to_datetime(df_rep['date'], errors='coerce')
                    df_rep = df_rep.dropna(subset=['date']).set_index('date')
                    final_rows = []
                    seq, total_working_days, absence_days, total_visits = 1, 0, 0, 0
                    
                    for d in date_range:
                        if d.weekday() == 4: continue
                        total_working_days += 1
                        day_name = days_ar[d.weekday()]
                        date_str = d.strftime('%Y-%m-%d')
                        
                        if d in df_rep.index:
                            row_df = df_rep.loc[d]
                            if isinstance(row_df, pd.DataFrame): row_df = row_df.iloc[0]
                            m_reps = row_df.get('mor_reports', 0) if pd.notna(row_df.get('mor_reports')) else 0
                            e_reps = row_df.get('eve_reports', 0) if pd.notna(row_df.get('eve_reports')) else 0
                            t_reps = row_df.get('tot_reports', 0) if pd.notna(row_df.get('tot_reports')) else 0
                            
                            m_start = clean_time(row_df.get('mor_start'))
                            m_end = clean_time(row_df.get('mor_end'))
                            e_start = clean_time(row_df.get('eve_start'))
                            e_end = clean_time(row_df.get('eve_end'))
                            
                            if m_reps == 0: m_start = 'لا يوجد زيارات'
                            if e_reps == 0: e_start = 'لا يوجد زيارات'
                            
                            total_visits += t_reps
                            final_rows.append({
                                'seq': seq, 'date': date_str, 'day': day_name,
                                'mor_reports': m_reps, 'eve_reports': e_reps, 'tot_reports': t_reps,
                                'mor_start': m_start, 'mor_end': m_end, 'eve_start': e_start, 'eve_end': e_end,
                                'total_time': clean_total_time(row_df.get('total_time')), 'notes': ''
                            })
                        else:
                            penalty = 0.5 if d.weekday() == 3 else 1.0
                            absence_days += penalty
                            final_rows.append({
                                'seq': seq, 'date': date_str, 'day': day_name,
                                'mor_reports': 0, 'eve_reports': 0, 'tot_reports': 0,
                                'mor_start': 'لا يوجد زيارات', 'mor_end': '', 'eve_start': 'لا يوجد زيارات', 'eve_end': '',
                                'total_time': '', 'notes': 'غياب'
                            })
                        seq += 1
                        
                    net_working_days = total_working_days - absence_days
                    average_visits = (total_visits / net_working_days) if net_working_days > 0 else 0
                    processed_reps[rep] = {
                        'rows': final_rows,
                        'summary': {'total_visits': total_visits, 'absence_days': absence_days, 'net_working_days': net_working_days, 'average_visits': average_visits}
                    }
                    total_all_visits += total_visits
                    total_all_absence += absence_days

                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                header_fill_main = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                summary_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                bold_font = Font(name="Arial", size=11, bold=True, color="000000")
                regular_font = Font(name="Arial", size=10, color="000000")
                absent_font = Font(name="Arial", size=10, bold=True, color="FF0000")
                border = Border(top=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"))
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for rep, data in processed_reps.items():
                    ws = wb.create_sheet(title=rep[:31])
                    ws.sheet_view.rightToLeft = True
                    
                    ws.merge_cells('A2:M2')
                    ws['A2'].value = f"مـــلخص الزيارات  اجمـــالي  بحسب الفترات\nفرع : {branch_name} | المندوب : {rep}    |   خلال الفترة من: {start_date.strftime('%Y/%m/%d')}  إلى: {end_date.strftime('%Y/%m/%d')}   |"
                    ws['A2'].font, ws['A2'].alignment = bold_font, center_align
                    ws.row_dimensions[2].height = 40
                    
                    headers_r4 = {'M': 'المندوب', 'L': 'تسلسل', 'K': 'التاريخ', 'J': 'اليوم', 'G': 'عدد التقارير', 'E': 'الفترة الصباحية', 'C': 'الفترة المسائية', 'B': 'إجمالي عدد ساعات\nودقائق العمل', 'A': 'ملاحظات'}
                    for col, text in headers_r4.items():
                        cell = ws[f"{col}4"]
                        cell.value, cell.font, cell.fill, cell.alignment, cell.border = text, bold_font, header_fill_main, center_align, border
                    for m in ['G4:I4', 'E4:F4', 'C4:D4', 'M4:M5', 'L4:L5', 'K4:K5', 'J4:J5', 'B4:B5', 'A4:A5']: ws.merge_cells(m)
                    
                    headers_r5 = {'I': 'الصباحية', 'H': 'المسائية', 'G': 'الاجمالي', 'F': 'وقت الجلسة الأولى', 'E': 'وقت الجلسة الأخيرة', 'D': 'وقت الجلسة الأولى', 'C': 'وقت الجلسة الأخيرة'}
                    for col, text in headers_r5.items():
                        cell = ws[f"{col}5"]
                        cell.value, cell.font, cell.fill, cell.alignment, cell.border = text, bold_font, header_fill_main, center_align, border
                    for r in ws['A4:M5']: 
                        for cell in r: cell.border = border
                    ws.row_dimensions[4].height, ws.row_dimensions[5].height = 30, 30
                    
                    rows = data['rows']
                    last_r = 5
                    for i, row_data in enumerate(rows):
                        r = i + 6
                        last_r = r
                        data_map = {'M': rep if i == 0 else "", 'L': row_data['seq'], 'K': row_data['date'], 'J': row_data['day'], 'I': row_data['mor_reports'], 'H': row_data['eve_reports'], 'G': row_data['tot_reports'], 'F': row_data['mor_start'], 'E': row_data['mor_end'], 'D': row_data['eve_start'], 'C': row_data['eve_end'], 'B': row_data['total_time'], 'A': row_data['notes']}
                        for col, val in data_map.items():
                            cell = ws[f"{col}{r}"]
                            cell.value, cell.border, cell.alignment = val, border, center_align
                            cell.font = absent_font if val == 'غياب' else regular_font

                    sum_start_r = last_r + 2
                    summary = data['summary']
                    sum_data = [
                        ("إجمالي عدد زيارات الأطباء:", summary['total_visits']),
                        ("إجمالي أيام الغياب (الخميس نصف يوم):", summary['absence_days']),
                        ("صافي أيام العمل:", summary['net_working_days']),
                        ("المتوسط اليومي للزيارات:", round(summary['average_visits'], 2))
                    ]
                    
                    for i, (label, val) in enumerate(sum_data):
                        curr_r = sum_start_r + i
                        ws.merge_cells(f'K{curr_r}:M{curr_r}')
                        label_cell = ws[f'K{curr_r}']
                        val_cell = ws[f'J{curr_r}']
                        label_cell.value, val_cell.value = label, val
                        for cell in [label_cell, val_cell]:
                            cell.font, cell.alignment, cell.fill, cell.border = bold_font, center_align, summary_fill, border
                        ws.merge_cells(f'I{curr_r}:J{curr_r}')
                        ws[f'I{curr_r}'].border = border
                            
                    widths = {'M': 25, 'L': 6, 'K': 12, 'J': 10, 'I': 10, 'H': 10, 'G': 10, 'F': 15, 'E': 15, 'D': 15, 'C': 15, 'B': 20, 'A': 15}
                    for col, w in widths.items(): ws.column_dimensions[col].width = w

                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.markdown('<div class="section-title">📊 ملخص نتائج الفرع والإنتاجية</div>', unsafe_allow_html=True)
                
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("اسم الفرع", branch_name)
                m2.metric("إجمالي المناديب", f"{len(processed_reps)} مناديب")
                m3.metric("مجموع الزيارات", f"{total_all_visits} زيارة")
                m4.metric("مجموع أيام الغياب", f"{total_all_absence} يوم")

                st.success(f"✅ تم الانتهاء بنجاح من معالجة بيانات فرع: {branch_name}")

                clean_branch_name = branch_name.replace(" ", "_")
                st.download_button(
                    label="📥 تحميل التقرير النهائي المنسق (Excel)",
                    data=output,
                    file_name=f"التقرير_النهائي_فرع_{clean_branch_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown('<div class="section-title">🔍 معاينة جداول حركة المناديب</div>', unsafe_allow_html=True)
                tabs = st.tabs([f"👤 {rep}" for rep in processed_reps.keys()])
                for tab, (rep_name, rep_info) in zip(tabs, processed_reps.items()):
                    with tab:
                        df_preview = pd.DataFrame(rep_info['rows'])
                        df_preview.columns = ['تسلسل', 'التاريخ', 'اليوم', 'زيارات صباحية', 'زيارات مسائية', 'إجمالي الزيارات', 'بداية الصباحية', 'نهاية الصباحية', 'بداية المسائية', 'نهاية المسائية', 'إجمالي الوقت', 'ملاحظات']
                        st.dataframe(df_preview, use_container_width=True)

            except Exception as e:
                st.error(f"حدث خطأ أثناء معالجة الملف: {e}")
    elif uploaded_file is None:
        st.info("💡 يرجى رفع ملف الإكسل الخام لتمكين زر إصدار التقرير.")


# --- 7. التحكم في معالجة التنقل والصفحات ---
if not st.session_state['logged_in']:
    show_login_page()
else:
    # الشريط الجانبي لتنقل النظام
    with st.sidebar:
        st.markdown(f"### 👤 مرحباً: {st.session_state['user_fullname']}")
        st.markdown("---")
        
        page = st.radio(
            "اختر الصفحة أو الأداة:",
            ["📊 نظام تقارير المناديب", "🔄 أداة التفقيط وتحويل الأرقام"],
            index=0
        )
        
        st.markdown("---")
        if st.button("🚪 تسجيل الخروج"):
            st.session_state['logged_in'] = False
            st.rerun()

    # التوجيه بحسب اختيار المستخدم
    if page == "📊 نظام تقارير المناديب":
        show_reports_page()
    elif page == "🔄 أداة التفقيط وتحويل الأرقام":
        components.html(HTML_TAFQEET_PAGE, height=1000, scrolling=True)
